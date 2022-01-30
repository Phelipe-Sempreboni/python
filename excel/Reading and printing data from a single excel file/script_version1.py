'''
Created by: Sempreboni.

Created in: 29/01/2022.

Last modification: 30/01/2022.
'''

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 1

# Importação de bibliotecas.
import os  # biblioteca que realiza operações com arquivos do sistema windows.
import openpyxl  # biblioteca para operações com arquivos do tipo excel.
import sys  # biblioteca que realiza operações nativas do sistema windows.

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 2

# Variável com o caminho do repositório de origem.
source_repository_path = r'C:\Users\pheli\OneDrive\Archives\2 - Github\Repositories\python\excel\Reading and printing data from a single excel file\archives'

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 3

# Variável com o nome do arquivo, que neste caso é fixo, mas pode ser alterado de acordo com a necessidade.
file_name = 'employee_registration.xlsx'

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 4

# Variável com a concatenação do repositório e nome do arquivo, para ter o caminho completo até a localização do arquivo.
# Notar que é utilizado duas barras, pois, somente uma barra represente quebra de texto, ou seja, mudança de linha sem sair do contexto atual.
file_repository_path = (source_repository_path + '\\' + file_name)

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 5

# Variável com o comando para verificar se o repositório origem, o mesmo da variável (source_repository_path) existe.
check_existence_source_repository = os.path.isdir(source_repository_path)

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 6

# Variável com o comando para verificar se o arquivo origem, o mesmo da variável (employees_file_name) existe.
check_existence_source_file = os.path.isfile(file_repository_path)

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 7

# Importante:
# Este bloco realiza a verificação se o repositório existe, ou seja, o repositório origem da variável (source_repository_path).
# Este bloco realiza a verificação se o arquivo existe, ou seja, o arquivo da variável (file_name).
# Caso o repositório ou o arquivo não exista, então haverá o erro e uma mensagem amigável será exibida ao usuário, tanto informando a inexistência do repositório ou do arquivo.
# Caso o repositório ou o arquivo não exista, então além da mensagem exibida ao usuário, o script será encerrado neste bloco 7 e não serão executados os demais scripts abaixo.

# Caso a cláusula seja verdadeira, então irá executar o comando abaixo.
# Verifica se o repositório origem existe. Caso existe, irá para o comando abaixo.
if check_existence_source_repository:

    # Caso a cláusula seja verdadeira, então irá executar o comando abaixo.
    # Verifica se o arquivo existe, caso exista, irá para o comando abaixo.
    if check_existence_source_file:

        # Variável com o comando para abertura do arquivo excel que desejamos trabalhar, onde é feito com a biblioteca (openpyxl).
        wbk_employee_registration = openpyxl.load_workbook(file_repository_path)

        # Continua com os demais comandos caso não haja erro no try, que neste caso, é na abertura do arquivo, indicando sua existência.
        pass

    # Caso a cláusula seja falsa, então irá executar o comando abaixo.
    # Se o arquivo não existir, será informado a mensagem abaixo ao usuário.
    else:

        # Mensagem ao usuário caso o repositório não exista.
        print('Arquivo (' + file_name + ') não existe. \nPor favor, verifique se o nome do arquivo está correto ou se existe no repositório origem e tente novamente. \nProcesso finalizado sem êxito.')

        # Caso ocorra o erro, então o script é encerrado e não prosseguirá para o próximo passo, assim evitando erros para o usuário posteriormente.
        sys.exit()

# Caso a cláusula seja falsa, então irá executar o comando abaixo.
# Se o repositório não existir, será informado a mensagem abaixo ao usuário.
else:

    # Mensagem ao usuário caso o repositório não exista.
    print('Repositório (' + source_repository_path + ') não existe. \nPor favor, verifique o caminho informado ou se o repositório existe e tente novamente. \nProcesso finalizado sem êxito.')

    # Caso ocorra o erro, então o script é encerrado e não prosseguirá para o próximo passo, assim evitando erros para o usuário posteriormente.
    sys.exit()

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 8

# Variável com o comando para verificar todas as planilhas disponiveis que temos para trabalhar na tabela.
wbk_employee_sheets_available = wbk_employee_registration.sheetnames

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 9

# Variável com o comando para escolher a planilha (sheet) que iremos querer trabalhar.
sheet_employee_registration = wbk_employee_registration['employee_registration']

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 10

# Bloco com try/except para para isolar um erro caso seja necessário, e, informar de uma maneira mais amigável ao usuário sobre um possível erro com uma, mensagem personalizada.
try:

    # O comando (iter_rows) realiza a iteração de comandos com as linhas.
    # O comando (iter_cols) realiza a iteração de comandos com as colunas.
    # Neste caso, estamos percorrendo a planilha, ou sheet, chamada ('employee_registration'), que está na variável (sheet_employee_registration).
    # Estamos percorrendo essa planilha, iniciando na linha minima 2, e lendo todos os valores.
    # Estamos utilizando o loop for para realizar a leitura completa dessa planilha, onde, para cada linha (row), retornamos os valores da instrução sheet_employee_registration.iter_rows(min_row=2, values_only=True).
    # Notar que teremos os valores devolvidos como tuplas, e tuplas são objetos imutáveis no python, ou seja, não é possível realizar nenhum tipo de modificação.
    for row in sheet_employee_registration.iter_rows(min_row=2, values_only=True):

        # Mensagem exibida ao usuário com os dados do loop for que foi executado.
        print(row)

# Caso ocorra um erro, então será executado o comando (except abaixo).
# Notar que é utilizado o (Exception) e renomeado para (erro), assim conseguimos capturar o erro e posteriormente exibir sua classe de erro juntamente com a mensagem, conforme o print abaixo.
except Exception as erro:

    # Mensagem exibida ao usuário caso ocorra de leitura dos dados da tabela em excel.
    print(f'Erro enquando era realizada a leitura dos dados. Tipo do erro: {erro.__class__}')

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #