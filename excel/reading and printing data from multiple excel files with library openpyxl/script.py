# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 1

# Importação de bibliotecas.
import os  # biblioteca que realiza operações com arquivos do sistema windows.
import openpyxl  # biblioteca para operações com arquivos do tipo excel.
import sys  # biblioteca que realiza operações nativas do sistema windows.

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 2

# Variável com o caminho do repositório de origem.
# Neste caso o caminho é fixo, ou seja, necessita de uma alteração manual.
#source_repository_path = r'C:\Users\pheli\OneDrive\Archives\2 - Github\Repositories\python\excel\Reading and printing data from a single excel file\archives'

# Neste caso o caminho é relativo, ou seja, não necessita de uma alteração manual.
# Este caminho, neste caso, seta o repositório que o este script está locado, logo, os arquivos necessitam estar junto do script no repositório.
path = os.path.abspath(__file__)
source_repository_path = os.path.dirname(path)
print('Repositório atual: ' + source_repository_path + '.')

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 3

# Variável com o nome da extensão do arquivo buscado, que neste caso é a extensão (xlsx).
# Notar que conseguimos verificar mais de um tipo de extensão do arquivo excel.
file_extension_one = 'xlsx'
file_extension_two = 'xlsm'

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 4

# Variável com o comando para verificar se o repositório origem, o mesmo da variável (source_repository_path) existe.
check_existence_source_repository = os.path.isdir(source_repository_path)

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Bloco 5

# Importante:
# Este bloco realiza a verificação se o repositório existe, ou seja, o repositório origem da variável (source_repository_path).
# Este bloco realiza a verificação se o arquivo que será lido é da extensão (xlsx ou xlsm), ou seja, um arquivo excel.
# Caso o repositório ou o arquivo não exista, então haverá o erro e uma mensagem amigável será exibida ao usuário, tanto informando a inexistência do repositório ou do arquivo.
# Caso o repositório ou o arquivo não exista, então além da mensagem exibida ao usuário, o script será encerrado neste bloco 6 e não serão executados os demais scripts abaixo.

# Declaração de variável global para utilização ao decorrer do script.
global wbk_employee_registration

# Caso a cláusula seja verdadeira, então irá executar o comando abaixo.
# Verifica se o repositório origem existe. Caso existe, irá para o comando abaixo.
if check_existence_source_repository:

    # loop for para verificar os arquivos existentes no repositório origem, ou seja, na variável (source_repository_path).
    for root, dicts, files in os.walk(source_repository_path):

        # loop for para verificar somente os arquivos, assim restringindo e melhorando a verificação.
        for file in files:

            # Caso a cláusula seja verdadeira, o comando abaixo é executado.
            # Este (if) realiza um slicing (fatiamento) do nome do arquivo que foi localizado pelos quatro últimos caracteres, que neste caso, seria o (xlsx ou xlsm) que procuramos.
            # Esse slicing (fatiamento) é comparado com a variável (file_extension_one e file_extension_two), que tem como valor (xlsx ou xlsm), ou seja, o valor que procuramos.
            if file[-4:] == file_extension_one or file[-4:] == file_extension_two:

                # Variável com o comando para abertura dos arquivos excel que desejamos trabalhar, onde é feito com a biblioteca (openpyxl).
                # Notar que neste caso, apontamos o caminho origem pela variável (source_repository_path) e concatenamos com a variável (file) que são os arquivos, visto que só teremos os arquivos no formato excel na extensão (xlsx ou xlsm) realizado pelo (if).
                # Para ficar mais visível, neste exemplo, a variável (file) iria trazer os arquivos do caminho origem nomeados como (employee_registration_one.xlsx, employee_registration_two.xlsx e employee_registration_macro.xlsm).
                # Temos um arquivo de extensão (txt) no caminho origem, porém, não será lido, pois, o if realizou a restrição neste caso.
                wbk_employee_registration = openpyxl.load_workbook(source_repository_path + '\\' + file)

                # Variável com o comando para verificar todas as planilhas disponiveis que temos para trabalhar na tabela.
                wbk_employee_sheets_available = wbk_employee_registration.sheetnames

                # Variável com o comando para escolher a planilha (sheet) que iremos querer trabalhar.
                sheet_employee_registration = wbk_employee_registration['employee_registration']

                # Bloco com try/except para para isolar um erro caso seja necessário, e, informar de uma maneira mais amigável ao usuário sobre um possível erro com uma, mensagem personalizada.
                try:

                    # O comando (iter_rows) realiza a iteração de comandos com as linhas.
                    # O comando (iter_cols) realiza a iteração de comandos com as colunas.
                    # Neste caso, estamos percorrendo a planilha, ou sheet, chamada ('employee_registration'), que está na variável (sheet_employee_registration).
                    # Estamos percorrendo essa planilha, iniciando na linha minima 2, e lendo todos os valores.
                    # Estamos utilizando o loop for para realizar a leitura completa dessa planilha, onde, para cada linha (row), retornamos os valores da instrução sheet_employee_registration.iter_rows(min_row=2, values_only=True).
                    # Notar que teremos os valores devolvidos como tuplas, e tuplas são objetos imutáveis no python, ou seja, não é possível realizar nenhum tipo de modificação.
                    for row in sheet_employee_registration.iter_rows(min_row=2, values_only=True):

                        # 1º maneira de impressão dos dados. Se for executar esse método, comente o 2º método.
                        # Mensagem exibida ao usuário com os dados do loop for que foi executado.
                        # Essa é uma mensagem simples. Abaixo temos um tipo de mensagem mais refinada.
                        print(row)

                        # 2º maneira de impressão dos dados. Se for executar esse método, comente o 1º método.
                        # Dessa maneira teremos os registros em formato de strings e não de um objeto lista. Sua utilização irá depender da intenção do usuário na utilização dos dados.
                        # coluna0 = row[0]  # matricula
                        # coluna1 = row[1]  # nome do funcionario
                        # coluna2 = row[2]  # data de nascimento
                        # coluna3 = row[3]  # idade
                        # coluna4 = row[4]  # data contratacao
                        # coluna5 = row[5]  # tipo da contratação
                        # coluna6 = row[6]  # situacao
                        # coluna7 = row[7]  # cargo
                        # coluna8 = row[8]  # diretoria
                        # coluna9 = row[9]  # setor
                        # coluna10 = row[10]  # salario
                        # coluna11 = row[11]  # data ultima atualizado dos registros

                        # Impressão dos dados.
                        # Assim ficaremos com um registro ao lado do outro.
                        # print(coluna0, coluna1, coluna2, coluna3, coluna4, coluna5, coluna6, coluna7, coluna8, coluna9, coluna10, coluna11)

                        # Impressão dos dados.
                        # Assim ficaremos com um registro ao lado do outro e com vírgulas.
                        # print(coluna0, ',', coluna1, ',', coluna2, ',', coluna3, ',', coluna4, ',', coluna5, ',', coluna6, ',', coluna7, ',', coluna8, ',', coluna9, ',', coluna10, ',', coluna11)

                        # Impressão dos dados.
                        # Assim ficaremos com um registro abaixo do outro.
                        # print(coluna0)
                        # print(coluna1)
                        # print(coluna2)
                        # print(coluna3)
                        # print(coluna4)
                        # print(coluna5)
                        # print(coluna6)
                        # print(coluna7)
                        # print(coluna8)
                        # print(coluna9)
                        # print(coluna10)
                        # print(coluna11)
                # Caso ocorra um erro, então será executado o comando (except abaixo).
                # Notar que é utilizado o (Exception) e renomeado para (erro), assim conseguimos capturar o erro e posteriormente exibir sua classe de erro juntamente com a mensagem, conforme o print abaixo.
                except Exception as erro:

                    # Mensagem exibida ao usuário caso ocorra de leitura dos dados da tabela em excel.
                    print(f'Erro enquando era realizada a leitura dos dados. Tipo do erro: {erro.__class__}. \nPor favor, verifique o caminho informado ou se o repositório existe e tente novamente. \nProcesso finalizado sem êxito.')

                    # Caso ocorra o erro, então o script é encerrado e não prosseguirá para o próximo passo, assim evitando erros para o usuário posteriormente.
                    sys.exit()

# Caso a cláusula seja falsa, então irá executar o comando abaixo.
# Se o repositório não existir, será informado a mensagem abaixo ao usuário.
else:

    # Mensagem ao usuário caso o repositório não exista.
    print('Repositório (' + source_repository_path + ') não existe. \nPor favor, verifique o caminho informado ou se o repositório existe e tente novamente. \nProcesso finalizado sem êxito.')

    # Caso ocorra o erro, então o script é encerrado e não prosseguirá para o próximo passo, assim evitando erros para o usuário posteriormente.
    sys.exit()

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #
