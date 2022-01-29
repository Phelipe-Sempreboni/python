'''
Created by: Sempreboni.

Created in: 29/01/2022.

Last modification: 29/01/2022.
'''

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Importing libraries
import openpyxl

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável com o caminho do repositório de origem.
source_repository_path = r'C:\Users\pheli\OneDrive\Archives\2 - Github\Repositories\python\projects\excel\reading a single excel file\archives'

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável com o nome do arquivo, que neste caso é fixo, mas pode ser alterado de acordo com a necessidade.
employees_file_name = 'employee_registration.xlsx'

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável com a concatenação do repositório e nome do arquivo, para ter o caminho completo até a localização do arquivo.
# Notar que é utilizado duas barras, pois, somente uma barra represente quebra de texto, ou seja, mudança de linha sem sair do contexto atual.
employee_file_repository_path = (source_repository_path + '\\' + employees_file_name)

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável com o comando para abertura do arquivo excel que desejamos trabalhar, onde é feito com a biblioteca (openpyxl).
wbk_employee = openpyxl.load_workbook(employee_file_repository_path)

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável com o comando para verificar todas as planilhas disponiveis que temos para trabalhar na tabela.
wbk_employee_sheets_available = wbk_employee.sheetnames

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável com o comando para escolher a planilha (sheet) que iremos querer trabalhar.
sheet_employee_registration = wbk_employee['employee_registration']

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# Variável de lista que está vazia, pois, irá receber os valores da planilha que ficarão armazenados e posteriormente adicionados no banco de dados.
# No final do script essa lista será esvaziada por questões de segurança, para não ocorrer erros e/ou acumulo de valores na próxima execução de script.
empty_list_to_receive_values = []

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #

# O comando (iter_rows) realiza a iteração de comandos com as linhas.
# O comando (iter_cols) realiza a iteração de comandos com as colunas.
# Neste caso, estamos percorrendo a planilha, ou sheet, chamada ('employee_registration'), que está na variável (sheet_employee_registration).
# Estamos percorrendo essa planilha, iniciando na linha minima 2, e lendo todos os valores.
# Estamos utilizando o loop for para realizar a leitura completa dessa planilha, onde, para cada linha (row), retornamos os valores da instrução sheet_employee_registration.iter_rows(min_row=2, values_only=True).
# Notar que teremos os valores devolvidos como tuplas, e tuplas são objetos imutáveis no python, ou seja, não é possível realizar nenhum tipo de modificação.
for row in sheet_employee_registration.iter_rows(min_row=2, values_only=True):

    empty_list_to_receive_values.append(row)

    tuples = tuple(empty_list_to_receive_values)

    print(tuples[0][0:2])

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------ #



